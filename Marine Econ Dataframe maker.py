import pandas as pd
from openpyxl import load_workbook
import os

'''Larger project was to calculate various economic indicators across sectors of the marine economy, this data was extremly 
crosssectional (each value was calculated according to jobs/gpd, province, typee of effect, year and the input industries.  This
meant that the final data took up several tables in several excel files.  The code below loops the various documents and
the tables within them to create a master dataset in "long" format which could be loaded into powerBI for easy visualization'''

if __name__ == '__main__':
    #excel files are messy and have multiple "tables" within each tab, which are also irregularly formated.  The code below
    #manually scrapes this particular format into a longform pandas data frame
    def topd(wb, ws):
        ws = wb[ws]
        dict = {}
        for column in range (1, 17):
            dictkey = ws.cell(row=4, column=column+1).value
            for row in range(35):
                cell = ws.cell(row=5+row, column=column+1)
                if dictkey in dict:
                    dict[dictkey].append(cell.value)
                else:
                    dict[dictkey] = [cell.value]

        dict['Type'] = []
        for row in range(35):
            dict['Type'].append('Direct')

        for column in range (1, 17):
            dictkey = ws.cell(row=4, column=column+1).value
            for row in range(35):
                cell = ws.cell(row=44+row, column=column+1)
                if dictkey in dict:
                    dict[dictkey].append(cell.value)
                else:
                    dict[dictkey] = [cell.value]
        for row in range(35):
            dict['Type'].append('Indirect')

        for column in range (1, 17):
            dictkey = ws.cell(row=4, column=column+1).value
            for row in range(35):
                cell = ws.cell(row=83+row, column=column+1)
                if dictkey in dict:
                    dict[dictkey].append(cell.value)
                else:
                    dict[dictkey] = [cell.value]
        for row in range(35):
            dict['Type'].append('Induced')

        return(pd.DataFrame(dict))

    #cucles through all tabs of each industry worksheet
    def cycle_ws(file):
        sheet_list = ['GDP, 2019', 'Jobs, 2019', 'GDP, 2018', 'Jobs, 2018']
        wb = load_workbook(file, data_only=True)
        df = topd(wb, sheet_list[0])
        df['Indicator'] = 'GDP'
        df['Year'] = 2019
        df2 = topd(wb, sheet_list[1])
        df2['Indicator'] = 'Jobs'
        df2['Year'] = 2019
        df3 = topd(wb, sheet_list[2])
        df3['Indicator'] = 'GDP'
        df3['Year'] = 2018
        df4 = topd(wb, sheet_list[3])
        df4['Indicator'] = 'Jobs'
        df4['Year'] = 2018

        df =df.append(df2)
        df = df.append(df3)
        df = df.append(df4)

        return df

    #cycles through each file in the folder
    def cycle_files(folder):
        fileList = os.listdir(folder)
        df = cycle_ws(folder + fileList[0])
        df['Industry'] = fileList[0]
        for x in range(1, len(fileList)):
            if fileList[x] != 'Long':
                df1 = cycle_ws(folder + fileList[x])
                df1['Industry'] = fileList[x]
                df = df.append(df1)
        return df



    df = cycle_files('C:/Marine Economy for POWER BI/')
    df.replace('.xlsx', '', inplace=True, regex=True)
    df.to_excel('C:/Marine Economy for POWER BI/Long/Longfile.xlsx')




